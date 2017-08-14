#encoding=utf-8

from openpyxl import Workbook, load_workbook
import random, os
# from common import ConfigManage
# from public_code.excel.ExcelManage import ExcelManage

class ExcelManage3(object):
    def __init__(self, filename, sheet_name = None):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.sheet_name = sheet_name
        if self.sheet_name == None:
            self.ws = self.wb.active
        else:
            self.ws = self.wb.get_sheet_by_name(self.sheet_name)

    def insert_data_by_col(self,start_row, col, data, new_file_name):
        u"""在指定列，插入相同数据"""
        for row in range(start_row, self.ws.max_row + 1):
            self.ws.cell(row=row, column=col).value = data
        self.wb.save(new_file_name)

    def insert_data_by_cell(self, row, col, data):
        u"""在指定单元格，插入数据"""
        self.ws.cell(row=row, column=col).value = data
        # self.wb.save(new_file_name)

    def save(self, new_file_name):
        self.wb.save(new_file_name)

    def get_datas(self):
        u"""获取excel表数据"""
        datas = []
        for row in range(1, self.ws.max_row + 1):
            data = []
            for col in range(1, self.ws.max_column + 1):
                if self.ws.cell(row=row, column=col).value is None:
                    data.append('')
                else:
                    data.append(self.ws.cell(row=row, column=col).value)
                    print self.ws.cell(row=row, column=col).value,
            print
            datas.append(data)
        return datas

if __name__ == '__main__':
    # em = ExcelManage3(os
    pass
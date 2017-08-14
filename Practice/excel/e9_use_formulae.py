from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# add a simple formula
ws["A1"] = "=SUM(1, 1)"
ws["A2"] = "=SUM(1, 1)"
ws["A3"] = "=SUM(A1,A2)"
wb.save(u"示例.xlsx")
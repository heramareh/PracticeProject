from openpyxl import Workbook
#未从文件系统生成真的excel文件，仅仅是实例化了一个Workbook实例
wb = Workbook()

#在excel的的最后位置加入一个sheet
ws1 = wb.create_sheet(u"光荣之路") 

#在在excel的第一个位置加入一个sheet
ws2 = wb.create_sheet("Mysheet", 0)

#修改sheet的名字
ws2.title=u"python excel 操作练习"

#设定sheet的背景样色
#颜色编码的网址：http://www.computerhope.com/htmcolor.htm
ws1.sheet_properties.tabColor = "FFFF00"
ws2.sheet_properties.tabColor = "FFA62F"

#通过名字获取某个sheet对象
ws3 = wb[u"python excel 操作练习"]

#打印sheet的名字
print ws3.title

#通过名字获取某个sheet对象
ws4 = wb.get_sheet_by_name(u"光荣之路")

#打印sheet的名字
print ws4.title

#打印所有的excel中的sheet名字，回用列表存储所有名字
print wb.sheetnames

#遍历输出excel中的所有sheet名字
for sheet in wb:
    print(sheet.title)


#生成真实的excel文件
wb.save(u"创建sheet.xlsx")
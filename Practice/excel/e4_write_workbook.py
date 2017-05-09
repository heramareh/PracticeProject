from openpyxl import Workbook

from openpyxl.utils import get_column_letter

#生成一个workbook对象
wb = Workbook()

#定义excel文件保存后的文件名
dest_filename = u'示例.xlsx'
 
#获取第一个sheet
ws1 = wb.active

#修改sheet的名字
ws1.title = u"新建的sheet1"
 
#
for row in range(1, 40):
    ws1.append(range(10))

ws2 = wb.create_sheet(title=u"新建的sheet2")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title=u"新建的sheet3")

for row in range(10, 20):
    for col in range(10, 20):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))  #get_column_letter函数可以把数字转换为列对应的字母


print(ws3['J10'].value)

wb.save(filename = dest_filename)
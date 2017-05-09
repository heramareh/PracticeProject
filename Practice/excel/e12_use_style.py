from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Border,GradientFill
from openpyxl.styles import colors
from openpyxl import Workbook
from copy import copy

wb = Workbook()
ws = wb.active

a1 = ws['A1']
d4 = ws['D4']

ft = Font(name=u"宋体",color=colors.RED,size=18,bold=True)
a1.font = ft
d4.font = ft

#a1.font.italic = True # is not allowed 

# If you want to change the color of a Font, you need to reassign it::
a1.font = Font(color="FFBB00",italic=True,bold=True) # the change only affects A1

a1.value=u"光荣之路"
d4.value=u"自动化测试"

#整行或者整列生效
col = ws.column_dimensions['E']   #E列会自动隐藏，需要手工取消隐藏
col.font = Font(bold=True)
row = ws.row_dimensions[1]
row.font = Font(underline="single")



#生成真实的excel文件
wb.save(u"示例.xlsx")
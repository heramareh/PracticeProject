import datetime
from openpyxl import Workbook

#guess_types表示类型推理，把字符串的内容进行相应的格式转换
wb = Workbook(guess_types=True)

ws = wb.active

# set date using a Python datetime
ws['A1'] = datetime.datetime(2010, 7, 21,10,9,8)

#excel被保存为自定义类型
print ws['A1'].number_format

print ws['A1'].value

# set percentage using a string followed by the percent sign
ws['B1'] = '3.14%'
print ws['B1'].value


ws['c1'] = '0.3'
print ws['c1'].value

print ws['c1'].number_format  #类型为General


ws['c2'] = 12
print ws['c2'].value

print ws['c2'].number_format  #类型为General


wb.save(u"示例.xlsx")  #注意：将直接覆盖，不是更新。
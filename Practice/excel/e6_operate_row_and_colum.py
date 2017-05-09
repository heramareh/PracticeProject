from openpyxl import Workbook
#未从文件系统生成真的excel文件，仅仅是实例化了一个Workbook实例
wb = Workbook()

#在excel的的最后位置加入一个sheet
ws = wb.create_sheet(u"光荣之路") 

#获取第一个sheet
ws=wb.active

#写入数据
ws["A1"]=1
ws["B1"]=2
ws["A2"]=u"光荣之路"
ws["B2"]="2017-2-19 19:19:19"


#获取所有行
print ws.rows

#获取某一行,例如获取第二行
print ws.rows[1]

#获取第二行的第二列单元格对象，并打印单元格的值
print ws.rows[1][1]
print ws.rows[1][1].value


#获取所有列
print ws.columns

#获取第二列的第二行单元格对象，并打印单元格的值
print ws.columns[1][1]
print ws.columns[1][1].value

#生成真实的excel文件，保存
wb.save(u"第一个文件.xlsx")

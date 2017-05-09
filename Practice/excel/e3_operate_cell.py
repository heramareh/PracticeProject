from openpyxl import Workbook
#未从文件系统生成真的excel文件，仅仅是实例化了一个Workbook实例
wb = Workbook()

#在excel的的最后位置加入一个sheet
ws = wb.create_sheet(u"光荣之路") 

#获取第一个sheet
ws=wb.active

#写入数据
ws["A1"]=1
ws["A2"]=u"中国"
ws["A3"]="2017-2-19 19:19:19"

#通过单元格字母标识，获取某个单元格对象
cell=ws["A2"]
#打印单元格的值
print cell.value,type(cell.value)   #中文和字母均
print ws["A1"].value,type(ws["A1"].value)  

#通过单元格序号，设定某个单元格对象的值
#单元格和列，均从1开始编号
cell=ws.cell(row=4, column=2, value=u"单元格")
print cell.value

#修改并读取一个单元格内容
cell.value=u"改了一下"
print cell.value


#创建100个单元格
for i in range(1,11):
    for j in range(1,11):
        ws.cell(row=i, column=j, value=str(i)+u"行"+str(j)+u"列")

#获取最大列和最大行
print ws.max_row
print ws.max_column

#获取所有单元格的值和坐标值
#for row in ws.iter_rows(range_string="A1:j10"):  #获取指定区间的所有单元格
for row in ws.iter_rows():
    for cell in row:
        print cell,cell.value,cell.coordinate



wb.save(u"第一个文件.xlsx")

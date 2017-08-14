from openpyxl import Workbook,load_workbook
import datetime

wb=load_workbook(u'第一个文件.xlsx')

#获取第一个sheet
ws=wb.active
ws['A1']="gloryroad"


wb.save(u"第一个文件.xlsx")  #注意：将直接覆盖，不是更新。

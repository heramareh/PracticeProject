#encoding-utf-8

from openpyxl import Workbook,load_workbook
import os

class ExcelManage(object):
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = Workbook()
        self.ws = self.wb.active

    def input_to_excel(self, data_dict, sheet=None):
        if sheet is not None:
            self.ws = self.wb.create_sheet(sheet)
        self.ws.append(data_dict.keys())
        self.ws.append(data_dict.values())
        self.ws.append()
        self.wb.save(self.filepath)


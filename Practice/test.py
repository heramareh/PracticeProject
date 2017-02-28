#encoding=utf-8

# import gloryroad

# if __name__ == '__main__':
# #     print gloryroad.fun111()
# #     print gloryroad.add(1,2)
# #     print gloryroad.sub(2,1)
#     try:
#         fp = open("d:\\test\\test.log",'r')
#         content = fp.readlines()
#     except IOError:
#         print "IOError!"
#     finally:
#         # 检测某个变量是否有定义的三种方法：
#         if vars().has_key('fp'):
# #         if 'fp' in dir():
# #         if 'fp' in locals().keys():
#             print "finally!"
#             fp.close()

'''写一个逐页显示文本文件的程序。
每次显示文本文件的指定行数，
暂停并向用户提示'按任意键继续'，按键后继续执行'''
# import os
# def get_content_by_lines(path,lines):
#     try:
#         if not os.path.exists(path):
#             raise IOError("path not exists!")
#         if not os.path.isfile(path):
#             raise IOError("path not a file!")
#         lineno = 0
#         with open(path) as fp:
#             for eachline in fp:
#                 if lineno % lines == 0 and lineno != 0:
#                     os.system('pause')
#                 print eachline,
#                 lineno += 1
#     except Exception,e:
#         print str(e)

# get_content_by_lines("d:\\test\\d.txt",5)

# import win32api,time
# import win32con
# # os.system('pause')
# while False:
#     print 1
#     # 回车
#     win32api.keybd_event(38,0,0,0)
#     win32api.keybd_event(38,0,win32con.KEYEVENTF_KEYUP,0)
#     time.sleep(0.1)
#     # 上方向键
#     win32api.keybd_event(13,0,0,0)
#     win32api.keybd_event(13,0,win32con.KEYEVENTF_KEYUP,0)
#     time.sleep(0.1)
#     print 2

# def foo():
#     print "calling foo()..."
#     aString = 'bar'
#     anInt = 123
#     print "foo()'s globals:",globals().keys()
#     print "foo()'s locals:",locals().keys()
# a = 100
# print "__main__'s globals:",globals().keys()
# print "__main__'s locals:",locals().keys()
# foo()

# class P(object):
#     def __init__(self):
#         print "P's constructor"
#         
# class C(P):
# #     pass
#     def __init__(self):
#         print "C's constructor"

# c = C()

class SortedKeyDict(dict):
    def keys(self):
        return sorted(super(SortedKeyDict,self).keys())

# d = SortedKeyDict((('a',1),('b',2),('c',3)))
# print [key for key in d]
# print d.keys()

# class A(object): pass
# class B(A): pass
# class C(A): pass
# class D(B, C): pass
# print D.__mro__

# 修改文件夹下的所有文件及文件夹的修改时间
import os
def recname(path):
    print "path:",path
    os.system("pause")
    if os.path.exists(path):
        for root,dirs,files in os.walk(path):
            for dir in dirs:
#                 print os.path.join(root,dir)
                os.utime(os.path.join(root,dir),None)
            for file in files:
#                 print os.path.join(root,file)
                os.utime(os.path.join(root,file),None)
        print "done!"
    else:
        print "path not exists!"
    os.system("pause")

# path = raw_input("please input path:")
# recname(path)

#修改文件夹下所有.java和.rftdef结尾的文件内容中指定字符串替换为新字符串
def change(path,old_str,new_str):
    print "path:",path
    os.system("pause")
    if os.path.exists(path):
        for root,dirs,files in os.walk(path):
            for file in files:
                if os.path.splitext(file)[1] in ['.java','.rftdef']:
                    file_path = os.path.join(root,file)
                    content = ''
                    with open(file_path,'rb') as fp:
                        content = fp.read()
                        content_list = content.split(old_str)
                        content = new_str.join(content_list)
                    with open(file_path,'wb') as fp:
                        fp.write(content)
        print "done!"
    else:
        print "path not exists!"
    os.system("pause")

#删除目录下所有指定的目录名及目录下的文件和文件夹
def deletedir(path,dirname):
    print "path:",path
    os.system("pause")
    if os.path.exists(path):
        for root,dirs,files in os.walk(path):
            for dir in dirs:
                if dir == dirname:
                    dir_path = os.path.join(root, dir)
                    print dir_path
                    for root1,dirs1,files1 in os.walk(dir_path):
                        print files1
                        for file1 in files1:
                            file_path = os.path.join(root1,file1)
                            os.remove(file_path)
                    for root2,dirs2,files2 in os.walk(dir_path,topdown = False):
                        for dir in dirs2:
                            try:
                                os.removedirs(os.path.join(root2,dir))
                            except:
                                continue
                    try:
                        os.removedirs(dir_path)
                    except:
                        continue
        print "done!"
    else:
        print "path not exists!"
    os.system("pause")

# path = raw_input("please input path:")
# deletedir(path,'jiuqi')

def x(n):
    global count
    count += 1
    if n<=3:
        return 1
    else:
        return x(n-2)+x(n-4)+1

# count = 0
# print x(8)
# print count

# import re
# p = re.compile('^.*[a-z]+$', re.I|re.S)
# string = """I am a Boy
#         Your a beautiful Girl
#         Right"""
# matchResult = p.search(string)
# if matchResult:
#     print matchResult.group()
# else:
#     print "no string found!"

import re

# s = 'aaa111aaa , bbb222 , 333ccc'
# # 指定前后肯定断言
# print re.findall( r'(?<=[a-z]{3})\d+(?=[a-z]+)', s) 
# 
# # 只指定后向肯定断言
# print re.findall( r'\w+\d+(?=[a-z]+)', s) 
# 
# # 只指定前向肯定断言
# print re.findall( r'(?<=[a-z]{3})\d+', s) 
# 
# # 普通匹配方法
# print re.findall (r'[a-z]+(\d+)[a-z]+', s)
# # 下面是一个错误的实例
# try:
#   matchResult = re.findall( r'(?<=[a-z]+)\d+(?=[a-z]+)', s) 
# except Exception, e:
#   print e
# else:
#   print matchResult
  
# s = 'aaa111aaa , bbb222 , 333ccc'
# # 指定前后否定断言，不满足前三个字母和后三个字母的条件
# print re.findall( r'(?<![a-z]{3})\d+(?![a-z]+)', s) 
# 
# # 只指定后向否定断言
# print re.findall( r'\w+\d+(?![a-z]+)', s) 
# 
# # 只指定前向否定断言
# print re.findall( r'(?<![a-z]{3})\d+', s) 
# 
# # 普通匹配方法
# print re.findall (r'[a-z]+(\d+)[a-z]+', s)
# # 下面是一个错误的实例
# try:
#   matchResult = re.findall( r'(?<![a-z]+)\d+(?![a-z]+)', s) 
# except Exception, e:
#   print e
# else:
#   print matchResult
  
import openpyxl
# wb = openpyxl.Workbook()
# ws = wb.active
# ws['A3'] = "hahaha"
# ws.append([1, "gloryroad", "中国"])
# ws.append([2, "gloryroad", "北京"])
# # wb.save(u"d:\\test\\excel\\gloryroad.xlsx")
# print dir(wb)

# wb = openpyxl.Workbook()
# ws1 = wb.create_sheet(u"光荣之路1")
# ws2 = wb.create_sheet(u"光荣之路2")
# ws3 = wb.create_sheet(u"光荣之路3")
# ws4 = wb.create_sheet(u"光荣之路4")
# ws1.sheet_properties.tabColor = "FFFF00"
# ws2.sheet_properties.tabColor = "3BB9FF"
# ws3.sheet_properties.tabColor = "7FFFD4"
# ws4.sheet_properties.tabColor = "F62217"
# for i in range(1,6):
#     for j in range(1,6):
#         ws1.cell(row=i,column=j,value=str(i)+','+str(j))
# 
# for row in ws1.iter_rows():
#     for cell in row:
#         print cell,cell.value,cell.coordinate
# 
# wb.save("d:\\test\\excel\\e2.xlsx")
import string
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
def create_excel(lista,content_list):
    wb = Workbook()
    ws = wb.active
    #边框
    #粗细
    thin1 = Side(border_style="thin", color="000000")
    #单双
    double1 = Side(border_style="double", color="000000")
    #上下左右边框
    border1 = Border(top=thin1, left=thin1, right=thin1, bottom=thin1)
    #字体、是否粗体、颜色、大小
    font1 = Font(name=u"微软雅黑",b=True, color="000000",size=12)
    #对齐方式
    al1 = Alignment(horizontal="center", vertical="center")
    #填充背景色
    fill1 = PatternFill("solid", fgColor="3090C7")
    border2 = Border(top=thin1, left=thin1, right=thin1, bottom=thin1)
    font2 = Font(name=u"微软雅黑",b=False, color="000000",size=12)
    al2 = Alignment(horizontal="center", vertical="center")
    fill2 = PatternFill("solid", fgColor="FFFF00")
    ws.append(lista)
    for i in content_list:
        ws.append(i)
    cols_num = string.ascii_uppercase[:len(lista)]
    for i in cols_num:
        cell = ws[i+"1"]
        cell.fill=fill1
        cell.border=border1
        cell.font=font1
        cell.alignment=al1
    for i in range(2,len(content_list)+2):
        for j in cols_num:
            cell = ws[j+str(i)]
            cell.fill=fill2
            cell.border=border2
            cell.font=font2
            cell.alignment=al2

    wb.save(u"d:\\test\\excel\\test.xlsx")

lista = ["序号","姓名","性别","学号","爱好","年龄"]
content = ["=row()-1","张三","男",1,"篮球",20]
content_list = []
for i in range(8):
    content_list.append(content)
print content_list
create_excel(lista,content_list)
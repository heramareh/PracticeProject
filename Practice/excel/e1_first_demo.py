from openpyxl import Workbook
#未从文件系统生成真的excel文件，仅仅是实例化了一个Workbook实例
wb = Workbook()

# Workbook()会创建至少sheet的excel对象，wb.active获取第一个sheet
ws = wb.active

# 直接在表格中赋值
ws['A1'] = 42

# 在表格中添加一行
ws.append([1, "china", "中国"])

# 在excel中写入日期对象，会自动转化为字符串
import datetime
print type(datetime.datetime.now())
ws['A3'] = datetime.datetime.now()

#生成真实的excel文件，保存
wb.save(u"d:\\test\\excel\\第一个文件.xlsx")
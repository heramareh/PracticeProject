from openpyxl import Workbook
import datetime
#未从文件系统生成真的excel文件，仅仅是实例化了一个Workbook实例
wb = Workbook(guess_types=True)

#在excel的的最后位置加入一个sheet
ws = wb.create_sheet(u"光荣之路") 

#获取第一个sheet
ws=wb.active

#写入数据
ws["A1"]='12%'
ws["A2"]=31.75
ws["A3"]=datetime.datetime.now()
ws["A4"]=u"中国"

#打印单元格，查看输出格式
print ws["A1"].value
print ws["A2"].value
print ws["A3"].value
print ws["A4"].value


wb.save(u"第一个文件.xlsx")  #注意：将直接覆盖，不是更新。
